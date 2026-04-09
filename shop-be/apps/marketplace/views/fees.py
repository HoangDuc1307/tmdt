"""
Thống kê phí sàn (Fees) - Theo dõi doanh thu và các khoản phí thu được từ giao dịch.
"""
from datetime import timedelta
import csv
import io
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from django.http import HttpResponse
from django.utils import timezone
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAdminUser
from rest_framework.response import Response

from ..models import AdminReportSnapshot, AdminAuditLog, PlatformFeeConfig
from ..revenue_utils import fee_statistics_payload, build_tx_day_map, combined_totals, top_transactions_mixed, current_order_fee_rate
from decimal import Decimal


@api_view(['POST'])
@permission_classes([IsAdminUser])
def save_fees_report(request):
    """Lưu lại snapshot báo cáo phí sàn để làm dữ liệu đối soát sau này."""
    stats_data = request.data.get('stats') or {}
    timeseries_data = request.data.get('timeseries') or {}
    
    # Cất vào database
    snapshot = AdminReportSnapshot.objects.create(
        report_type='FEES',
        snapshot_data={'stats': stats_data, 'timeseries': timeseries_data},
        created_by=request.user,
    )

    # Note lại hành động của admin
    AdminAuditLog.objects.create(
        admin=request.user,
        action='SAVE_FEES_REPORT',
        details="Đã lưu snapshot báo cáo phí sàn.",
        target_model="AdminReportSnapshot",
        target_id=str(snapshot.id)
    )

    return Response({'status': 'saved', 'message': 'Đã lưu snapshot báo cáo phí sàn.'})


@api_view(['GET'])
@permission_classes([IsAdminUser])
def fee_statistics(request):
    """Lấy tổng hợp số liệu về doanh thu, phí sàn và số lượng giao dịch (Transaction + Order đã thanh toán)."""
    days = int(request.query_params.get('days', 7))
    if days < 1:
        days = 7
    if days > 90:
        days = 90

    payload = fee_statistics_payload(days)
    payload['order_fee_percent'] = float((current_order_fee_rate() * Decimal('100')).quantize(Decimal('0.01')))
    return Response(payload)


@api_view(['GET', 'PUT'])
@permission_classes([IsAdminUser])
def fee_config(request):
    """
    Xem/chỉnh % phí sàn áp dụng cho Order.
    """
    config = PlatformFeeConfig.get_solo()

    if request.method == 'GET':
        return Response({
            'fee_percent': float(config.fee_percent),
            'updated_at': config.updated_at,
            'updated_by': config.updated_by.username if config.updated_by else None,
        })

    raw_percent = request.data.get('fee_percent')
    try:
        fee_percent = Decimal(str(raw_percent))
    except Exception:
        return Response({'error': 'fee_percent không hợp lệ'}, status=400)

    if fee_percent < 0 or fee_percent > 100:
        return Response({'error': 'fee_percent phải trong khoảng 0-100'}, status=400)

    config.fee_percent = fee_percent.quantize(Decimal('0.01'))
    config.updated_by = request.user
    config.save(update_fields=['fee_percent', 'updated_by', 'updated_at'])

    AdminAuditLog.objects.create(
        admin=request.user,
        action='UPDATE_PLATFORM_FEE_PERCENT',
        details=f"Cập nhật phí sàn cho Order: {config.fee_percent}%",
        target_model='PlatformFeeConfig',
        target_id=str(config.id),
    )
    return Response({
        'message': 'Cập nhật % phí sàn thành công',
        'fee_percent': float(config.fee_percent),
    })


@api_view(['GET'])
@permission_classes([IsAdminUser])
def fee_top_transactions(request):
    """5 giao dịch phí cao nhất (marketplace Transaction + Order đã thanh toán)."""
    return Response(top_transactions_mixed(5))


@api_view(['GET'])
@permission_classes([IsAdminUser])
def export_fees_report_csv(request):
    """
    Xuất file Excel báo cáo phí sàn. Show cả tổng quan, chi tiết theo ngày và top giao dịch.
    """
    days = int(request.query_params.get('days', 7))
    if days < 1:
        days = 7
    if days > 90:
        days = 90

    now = timezone.now()
    since = now - timedelta(days=days)
    today = timezone.localdate()
    start_date = today - timedelta(days=days - 1)

    summary_dict = combined_totals()
    last_dict = combined_totals(since_dt=since)
    rev_n = float(last_dict['total_revenue'].quantize(Decimal('0.01')))
    fee_n = float(last_dict['total_platform_fee'].quantize(Decimal('0.01')))
    cnt = summary_dict['total_transactions']
    fee_total = summary_dict['total_platform_fee']
    avg_fee = float((fee_total / Decimal(cnt)).quantize(Decimal('0.01'))) if cnt else 0.0
    summary = {
        'total_revenue': summary_dict['total_revenue'],
        'total_platform_fee': summary_dict['total_platform_fee'],
        'total_transactions': summary_dict['total_transactions'],
    }

    labels = [(start_date + timedelta(days=i)).isoformat() for i in range(days)]
    tx_map_full = build_tx_day_map(start_date, include_orders=True)
    tx_map = {k: {'revenue': v['revenue'], 'fee': v['fee']} for k, v in tx_map_full.items()}

    top_rows = top_transactions_mixed(5)

    filename = f"fees-report-{today.isoformat()}"
    
    # Khởi tạo file Excel bằng openpyxl
    wb = Workbook()
    ws = wb.active
    ws.title = "Fees Report"

    # Định dạng font cho tiêu đề
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True)

    # PHẦN 1: Bảng tổng quan (Summary)
    ws.append(['BÁO CÁO PHÍ SÀN'])
    ws['A1'].font = title_font
    ws.append(['Số ngày', days])
    ws.append([])
    ws.append(['Chỉ số', 'Giá trị'])
    # In đậm hàng header
    for cell in ws[4]:
        cell.font = header_font

    ws.append(['Tổng doanh thu', float(summary['total_revenue'] or 0)])
    ws.append(['Tổng phí sàn', float(summary['total_platform_fee'] or 0)])
    ws.append(['Tổng giao dịch', summary['total_transactions'] or 0])
    ws.append([f'Doanh thu {days} ngày gần nhất', rev_n])
    ws.append([f'Phí sàn {days} ngày gần nhất', fee_n])
    ws.append(['Phí trung bình mỗi giao dịch', avg_fee])

    # PHẦN 2: Dữ liệu chi tiết từng ngày (Timeseries)
    ws.append([])
    ws.append(['DỮ LIỆU THEO NGÀY'])
    ws.cell(row=ws.max_row, column=1).font = title_font
    ws.append(['Ngày', 'Doanh thu', 'Phí sàn'])
    for cell in ws[ws.max_row]:
        cell.font = header_font

    for label in labels:
        tx_data = tx_map.get(label, {})
        ws.append([
            label,
            tx_data.get('revenue', 0),
            tx_data.get('fee', 0),
        ])

    # PHẦN 3: Danh sách 5 giao dịch có phí sàn cao nhất
    ws.append([])
    ws.append(['TOP 5 GIAO DỊCH CÓ PHÍ CAO NHẤT'])
    ws.cell(row=ws.max_row, column=1).font = title_font
    ws.append(['ID', 'Bài đăng', 'Người mua', 'Số tiền', 'Phí sàn', 'Ngày'])
    for cell in ws[ws.max_row]:
        cell.font = header_font

    for row in top_rows:
        ws.append([
            row['id'],
            row['listing_title'],
            row['buyer_username'],
            float(row['amount']),
            float(row['platform_fee']),
            row['created_at'],
        ])

    # Tự động chỉnh độ rộng cột nhìn cho đẹp
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    wb.save(response)
    
    return response
